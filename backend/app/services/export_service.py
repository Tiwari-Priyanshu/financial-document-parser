"""
Report generation: Excel, PDF and CSV.

All three formats are driven by the same parser specs, so a field added to a
DocumentSpec appears in every export automatically - there is no per-format
column list to keep in sync.

Exports always use `effective_data` (manual corrections winning over raw AI
output), which is what the spec means by "exported reports should always
contain the latest approved data".

Files are generated in memory and streamed. Writing them to disk first would
mean cleaning up temp files and would not survive Render's ephemeral
filesystem across restarts.
"""

import csv
import io
import logging
from datetime import datetime, timezone
from typing import Any, Optional

from app.models.document import Document
from app.models.enums import DocumentType
from app.models.report import ParsedReport
from app.parsers.base import FieldType
from app.parsers.registry import get_spec

logger = logging.getLogger(__name__)


class ExportError(Exception):
    """Raised when a report cannot be generated."""


def _format_value(value: Any, field_type: Optional[FieldType] = None) -> str:
    """Render a parsed value for display in a report cell."""
    if value is None or value == "":
        return "-"
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)) and field_type == FieldType.NUMBER:
        return f"{value:,.2f}"
    if isinstance(value, list):
        return f"{len(value)} item(s)"
    if isinstance(value, dict):
        return "; ".join(f"{k}={v}" for k, v in value.items())
    return str(value)


def _field_rows(document: Document, report: ParsedReport) -> list[dict[str, str]]:
    """
    Build the extracted-fields table, one row per field in the spec.

    Rows come from the spec rather than from the data, so a field the AI failed
    to find still appears - shown as "-" with its validation message. A missing
    field is information; silently omitting it is not.
    """
    spec = get_spec(document.document_type) if document.document_type else None
    if spec is None:
        return [
            {"field": key, "value": _format_value(value), "status": "n/a", "note": ""}
            for key, value in (report.effective_data or {}).items()
        ]

    issues_by_field: dict[str, list[dict]] = {}
    for issue in report.validation_errors or []:
        issues_by_field.setdefault(issue.get("field", ""), []).append(issue)

    data = report.effective_data or {}
    corrected_keys = set((report.corrected_data or {}).keys())
    original = report.parsed_data or {}

    rows = []
    for field in spec.fields:
        value = data.get(field.name)
        issues = issues_by_field.get(field.name, [])

        if issues:
            status = "Error" if any(
                i.get("severity") == "error" for i in issues
            ) else "Warning"
            note = "; ".join(i.get("message", "") for i in issues)
        elif value in (None, "", []):
            status = "Missing" if field.mandatory else "Not present"
            note = "Required field not found" if field.mandatory else ""
        else:
            status = "Valid"
            note = ""

        # Flag values a human changed, so a reviewer can see what was touched.
        if field.name in corrected_keys and original.get(field.name) != value:
            note = (
                f"Manually corrected from '{_format_value(original.get(field.name))}'"
                + (f". {note}" if note else "")
            )

        rows.append({
            "field": field.label,
            "value": _format_value(value, field.type),
            "status": status,
            "note": note,
        })

    return rows


def _metadata_rows(document: Document, report: ParsedReport) -> list[tuple[str, str]]:
    """The header block every export carries, per the spec's report contents."""
    return [
        ("Document Name", document.document_name),
        ("Document Type", (document.document_type or DocumentType.UNKNOWN).value
                          .replace("_", " ").title()),
        ("File Size", f"{document.file_size / 1024:.1f} KB"),
        ("Uploaded By", f"{document.uploader_name} ({document.uploader_email})"),
        ("Upload Date", document.created_at.strftime("%d %b %Y, %H:%M UTC")),
        ("Processing Status", document.status.value.replace("_", " ").title()),
        ("Processing Time", f"{document.processing_time:.2f} s"
                            if document.processing_time else "-"),
        ("Extraction Method", (report.extraction_method or "-")
                              .replace("_", " ").title()),
        ("AI Confidence", f"{report.confidence_score:.0%}"
                          if report.confidence_score is not None else "-"),
        ("Validation Status", report.validation_status.value.title()),
        ("Review Status", report.review_status.value.replace("_", " ").title()),
        ("Reviewed By", report.reviewer_name or "-"),
        ("Remarks", report.remarks or "-"),
        ("Report Generated", datetime.now(timezone.utc)
                             .strftime("%d %b %Y, %H:%M UTC")),
    ]


# --- CSV ----------------------------------------------------------------


def generate_csv(document: Document, report: ParsedReport) -> bytes:
    """
    Flat CSV: metadata block, then the extracted fields table.

    utf-8-sig rather than plain utf-8 - the BOM makes Excel on Windows read
    rupee symbols and accented names correctly instead of showing mojibake.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow(["FINANCIAL DOCUMENT PARSING REPORT"])
    writer.writerow([])
    for label, value in _metadata_rows(document, report):
        writer.writerow([label, value])

    writer.writerow([])
    writer.writerow(["EXTRACTED FIELDS"])
    writer.writerow(["Field", "Value", "Validation", "Notes"])
    for row in _field_rows(document, report):
        writer.writerow([row["field"], row["value"], row["status"], row["note"]])

    # Line items and transactions are nested arrays - give them their own
    # section rather than collapsing them to "2 item(s)".
    data = report.effective_data or {}
    for key in ("line_items", "transactions"):
        items = data.get(key)
        if isinstance(items, list) and items and isinstance(items[0], dict):
            writer.writerow([])
            writer.writerow([key.replace("_", " ").upper()])
            headers = list(items[0].keys())
            writer.writerow([h.replace("_", " ").title() for h in headers])
            for item in items:
                writer.writerow([_format_value(item.get(h)) for h in headers])

    return buffer.getvalue().encode("utf-8-sig")


# --- Excel --------------------------------------------------------------


def generate_excel(document: Document, report: ParsedReport) -> bytes:
    """
    Multi-sheet workbook: Summary, Extracted Fields, Validation, plus a sheet
    per nested array.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ExportError("openpyxl is not installed") from exc

    wb = Workbook()

    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14, color="1F3864")
    label_font = Font(bold=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_fills = {
        "Valid": PatternFill("solid", start_color="E2EFDA"),
        "Warning": PatternFill("solid", start_color="FFF2CC"),
        "Error": PatternFill("solid", start_color="FCE4E4"),
        "Missing": PatternFill("solid", start_color="FCE4E4"),
    }

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Financial Document Parsing Report"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")

    row = 3
    for label, value in _metadata_rows(document, report):
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 55

    # --- Extracted fields sheet ---
    ws2 = wb.create_sheet("Extracted Fields")
    headers = ["Field", "Value", "Validation", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for r, field_row in enumerate(_field_rows(document, report), start=2):
        values = [field_row["field"], field_row["value"],
                  field_row["status"], field_row["note"]]
        for c, value in enumerate(values, start=1):
            cell = ws2.cell(row=r, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 4))
        fill = status_fills.get(field_row["status"])
        if fill:
            ws2.cell(row=r, column=3).fill = fill

    for col, width in zip("ABCD", (26, 34, 14, 60)):
        ws2.column_dimensions[col].width = width
    ws2.freeze_panes = "A2"   # keep headers visible when scrolling

    # --- Validation sheet ---
    ws3 = wb.create_sheet("Validation")
    for col, header in enumerate(["Field", "Rule", "Severity", "Message"], start=1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    issues = report.validation_errors or []
    if issues:
        for r, issue in enumerate(issues, start=2):
            ws3.cell(row=r, column=1, value=issue.get("field", ""))
            ws3.cell(row=r, column=2, value=issue.get("rule", ""))
            ws3.cell(row=r, column=3, value=issue.get("severity", ""))
            ws3.cell(row=r, column=4, value=issue.get("message", ""))
    else:
        ws3.cell(row=2, column=1, value="No validation issues found")

    for col, width in zip("ABCD", (24, 28, 12, 80)):
        ws3.column_dimensions[col].width = width
    ws3.freeze_panes = "A2"

    # --- Nested arrays get their own sheets ---
    data = report.effective_data or {}
    for key in ("line_items", "transactions"):
        items = data.get(key)
        if not (isinstance(items, list) and items and isinstance(items[0], dict)):
            continue
        sheet = wb.create_sheet(key.replace("_", " ").title()[:31])
        keys = list(items[0].keys())
        for col, k in enumerate(keys, start=1):
            cell = sheet.cell(row=1, column=col, value=k.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
        for r, item in enumerate(items, start=2):
            for c, k in enumerate(keys, start=1):
                value = item.get(k)
                sheet.cell(row=r, column=c,
                           value=value if isinstance(value, (int, float))
                           else _format_value(value))
        for col in range(1, len(keys) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 22
        sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- PDF ----------------------------------------------------------------


def generate_pdf(document: Document, report: ParsedReport) -> bytes:
    """
    A4 report with a metadata block, the extracted fields table, and a
    validation section. Uses ReportLab's Platypus flowables so tables split
    across pages correctly instead of being clipped.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError as exc:
        raise ExportError("reportlab is not installed") from exc

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
        title=f"Parsing Report - {document.document_name}",
    )

    styles = getSampleStyleSheet()
    navy = colors.HexColor("#1F3864")
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Heading1"],
        fontSize=17, textColor=navy, spaceAfter=4, alignment=TA_LEFT,
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"],
        fontSize=12, textColor=navy, spaceBefore=14, spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        "Cell", parent=styles["BodyText"], fontSize=8, leading=10,
    )

    story: list[Any] = []
    story.append(Paragraph("Financial Document Parsing Report", title_style))
    story.append(Paragraph(
        f"Generated {datetime.now(timezone.utc).strftime('%d %B %Y, %H:%M UTC')}",
        styles["Italic"],
    ))
    story.append(Spacer(1, 8))

    # --- Metadata ---
    story.append(Paragraph("Document Information", section_style))
    meta_data = [[Paragraph(f"<b>{label}</b>", cell_style),
                  Paragraph(str(value), cell_style)]
                 for label, value in _metadata_rows(document, report)]
    meta_table = Table(meta_data, colWidths=[45 * mm, 125 * mm])
    meta_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D0D0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F2F4F8")),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(meta_table)

    # --- Extracted fields ---
    story.append(Paragraph("Extracted Fields", section_style))
    rows = _field_rows(document, report)
    field_data = [[Paragraph(f"<b>{h}</b>", cell_style)
                   for h in ("Field", "Value", "Validation", "Notes")]]
    for row in rows:
        field_data.append([
            Paragraph(row["field"], cell_style),
            Paragraph(row["value"], cell_style),
            Paragraph(row["status"], cell_style),
            Paragraph(row["note"], cell_style),
        ])

    field_table = Table(
        field_data, colWidths=[38 * mm, 45 * mm, 20 * mm, 67 * mm],
        repeatRows=1,   # repeat the header on every page
    )
    style_commands = [
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D0D0")),
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for index, row in enumerate(rows, start=1):
        if row["status"] in ("Error", "Missing"):
            style_commands.append(
                ("BACKGROUND", (0, index), (-1, index), colors.HexColor("#FCE4E4")))
        elif row["status"] == "Warning":
            style_commands.append(
                ("BACKGROUND", (0, index), (-1, index), colors.HexColor("#FFF6DA")))
    field_table.setStyle(TableStyle(style_commands))
    story.append(field_table)

    # --- Validation ---
    story.append(Paragraph("Validation Results", section_style))
    issues = report.validation_errors or []
    if issues:
        issue_data = [[Paragraph(f"<b>{h}</b>", cell_style)
                       for h in ("Field", "Severity", "Message")]]
        for issue in issues:
            issue_data.append([
                Paragraph(issue.get("field", ""), cell_style),
                Paragraph(issue.get("severity", ""), cell_style),
                Paragraph(issue.get("message", ""), cell_style),
            ])
        issue_table = Table(
            issue_data, colWidths=[38 * mm, 22 * mm, 110 * mm], repeatRows=1)
        issue_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D0D0")),
            ("BACKGROUND", (0, 0), (-1, 0), navy),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(issue_table)
    else:
        story.append(Paragraph("All validation checks passed.", styles["BodyText"]))

    # --- Nested arrays ---
    data = report.effective_data or {}
    for key in ("line_items", "transactions"):
        items = data.get(key)
        if not (isinstance(items, list) and items and isinstance(items[0], dict)):
            continue
        story.append(PageBreak())
        story.append(Paragraph(key.replace("_", " ").title(), section_style))
        keys = list(items[0].keys())
        table_data = [[Paragraph(f"<b>{k.replace('_', ' ').title()}</b>", cell_style)
                       for k in keys]]
        for item in items:
            table_data.append(
                [Paragraph(_format_value(item.get(k)), cell_style) for k in keys])
        width = 170 * mm / max(len(keys), 1)
        nested = Table(table_data, colWidths=[width] * len(keys), repeatRows=1)
        nested.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D0D0")),
            ("BACKGROUND", (0, 0), (-1, 0), navy),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(nested)

    def footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#808080"))
        canvas.drawString(18 * mm, 10 * mm,
                          f"{document.document_name}  |  AI Financial Document Parser")
        canvas.drawRightString(A4[0] - 18 * mm, 10 * mm, f"Page {doc_.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return buffer.getvalue()


def safe_filename(document: Document, extension: str) -> str:
    """Build a download filename that is safe on every OS."""
    base = "".join(
        c if c.isalnum() or c in "-_" else "_"
        for c in document.document_name.rsplit(".", 1)[0]
    )[:60]
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    return f"{base}_report_{stamp}.{extension}"
